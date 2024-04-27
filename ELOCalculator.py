import openpyxl
import math
workbook = openpyxl.load_workbook('POKER SKORES.xlsx')
sheet = workbook.active
player_names = []
elo_ratings = []
for cell in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1, values_only=True):
    if cell[0] is None:
        break
    player_names.append(cell[0])
for cell in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2, values_only=True):
    if cell[0] is None:
        break
    elo_ratings.append(cell[0])
player_elo_dict = dict(zip(player_names, elo_ratings))
K = 100
for col in range(3, sheet.max_column + 1):
    print(f"Match {col - 2}:")
    num_players = sheet.cell(row=1, column=col).value
    total_buyins = sum(1 for row in range(2, sheet.max_row + 1) if sheet.cell(row=row, column=col).value is not None and '.' in str(sheet.cell(row=row, column=col).value))
    num_players += total_buyins
    participating_players = [sheet.cell(row=row, column=1).value for row in range(2, sheet.max_row + 1) if sheet.cell(row=row, column=col).value is not None]
    expected_win_pct = {}
    for player_name in participating_players:
        if player_name in player_elo_dict:
            player_elo = player_elo_dict[player_name]
            opponents_elo = [player_elo_dict[name] for name in participating_players if name != player_name and name in player_elo_dict]
            total_elo = player_elo + sum(opponents_elo)
            expected_win_pct[player_name] = player_elo / total_elo
    for player_name, e_score in expected_win_pct.items():
        print(f"{player_name}: E = {e_score:.3f}")
    print()
    for row in range(2, sheet.max_row + 1):
        player_name = sheet.cell(row=row, column=1).value
        
        if player_name is None:
            break
        player_rank = sheet.cell(row=row, column=col).value
        
        if player_rank is not None:
            positions = str(player_rank).replace('.', ' ').split()
            s_scores = []
            for pos in positions:
                if pos.strip():
                    rank = int(pos)
                    score_range = math.floor(num_players / 2)
                    
                    if 1 <= rank <= score_range:
                        s_score = 1 / rank
                        s_scores.append(s_score)
                    else:
                        s_scores.append(0)
            if s_scores:
                print(f"{player_name}: {', '.join(map(lambda x: f'{x:.3f}', s_scores))}")
                for s_score in s_scores:
                    rpre = player_elo_dict[player_name]
                    e_score = expected_win_pct[player_name]
                    rnew = rpre + K * (s_score - e_score)
                    player_elo_dict[player_name] = rnew
                    print(f"{player_name}: Rnew = {rnew:.2f}")
    print()
print(player_elo_dict)
row = 2
for player, elo in player_elo_dict.items():
    rounded_elo = round(elo)
    sheet.cell(row=row, column=2, value=rounded_elo)
    row += 1
workbook.save('POKER SKORES.xlsx')